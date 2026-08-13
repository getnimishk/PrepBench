"""
roadmap_import_service.py

Turns a roadmap file into a reviewable preview, then commits the approved
preview as a roadmap tree.

The governing constraint is that this must be **generic**. The reference
workbook's sheets are called "Kafka Master Syllabus" and "Journey Progress
Tracker", and matching on names like those would quietly make the feature
Kafka-only. Sheets are therefore classified by the *shape of their header
row*, so any workbook with a Phase/Topic/Objective-shaped table imports.
"""

import io
import json
import re
from datetime import datetime
from typing import List, Optional, Tuple, Dict, Any

import openpyxl
import pandas as pd
from sqlalchemy.orm import Session

from app.core.exceptions import ImportValidationException
from app.core.logging_config import logger
from app.models.roadmap import Roadmap, RoadmapPhase, RoadmapTopic, RoadmapResource, RoadmapTopicStatus
from app.schemas.roadmap import (
    RoadmapImportPreview, RoadmapImportTopic, RoadmapImportResource,
    RoadmapImportConfirm, RoadmapImportResult,
)
from app.services.roadmap_service import RoadmapService

# Header tokens that identify a curriculum table vs. a progress table. Scored
# rather than matched exactly, because a progress sheet also mentions "Topic"
# and "Phase" -- whichever signature scores higher wins.
SYLLABUS_TOKENS = ("phase", "module", "section", "topic", "title", "skill",
                   "objective", "goal", "success", "criteria", "outcome",
                   "hour", "effort", "estimate")
TRACKER_TOKENS = ("status", "progress", "start date", "started",
                  "completion", "completed", "finish", "evidence", "notes")

DEFAULT_PHASE_NAME = "General"

# Spreadsheets routinely end a table with a totals line ("TOTAL ESTIMATED
# HOURS | 134"). Imported naively that becomes a 134-hour topic in a phantom
# phase, which then corrupts every hours figure derived from it.
SUMMARY_ROW_PATTERN = re.compile(
    r"^\s*(total|totals|grand\s+total|sub\s*total|subtotal|sum|overall|summary)\b",
    re.IGNORECASE,
)

STATUS_ALIASES = {
    "not started": RoadmapTopicStatus.NOT_STARTED,
    "notstarted": RoadmapTopicStatus.NOT_STARTED,
    "todo": RoadmapTopicStatus.NOT_STARTED,
    "to do": RoadmapTopicStatus.NOT_STARTED,
    "pending": RoadmapTopicStatus.NOT_STARTED,
    "in progress": RoadmapTopicStatus.IN_PROGRESS,
    "inprogress": RoadmapTopicStatus.IN_PROGRESS,
    "doing": RoadmapTopicStatus.IN_PROGRESS,
    "started": RoadmapTopicStatus.IN_PROGRESS,
    "wip": RoadmapTopicStatus.IN_PROGRESS,
    "completed": RoadmapTopicStatus.COMPLETED,
    "complete": RoadmapTopicStatus.COMPLETED,
    "done": RoadmapTopicStatus.COMPLETED,
    "finished": RoadmapTopicStatus.COMPLETED,
    "skipped": RoadmapTopicStatus.SKIPPED,
    "skip": RoadmapTopicStatus.SKIPPED,
    "n/a": RoadmapTopicStatus.SKIPPED,
    "na": RoadmapTopicStatus.SKIPPED,
}


def _clean(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_status(value: Any) -> RoadmapTopicStatus:
    text = _clean(value)
    if not text:
        return RoadmapTopicStatus.NOT_STARTED
    # Separators normalised so a spreadsheet's "In Progress", a JSON export's
    # "in_progress", and a hyphenated "in-progress" all land on one alias.
    key = re.sub(r"\s+", " ", re.sub(r"[_\-]+", " ", text.strip().lower()))
    return STATUS_ALIASES.get(key, RoadmapTopicStatus.NOT_STARTED)


def _parse_progress(value: Any) -> int:
    """
    Percent, clamped to 0-100.

    A float strictly between 0 and 1 is read as a fraction, because that is how
    Excel stores a percent-formatted cell (40% is on disk as 0.4). Whole
    numbers are taken at face value. Anything genuinely ambiguous survives into
    the staged review, where it is visible before anything is written.
    """
    if value is None:
        return 0
    try:
        number = float(value)
    except (TypeError, ValueError):
        return 0
    if 0 < number < 1:
        number *= 100
    return int(max(0, min(100, round(number))))


def _parse_hours(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        hours = float(value)
    except (TypeError, ValueError):
        # "3-4h" and similar -- take the first number rather than dropping it.
        match = re.search(r"\d+(?:\.\d+)?", str(value))
        if not match:
            return None
        hours = float(match.group())
    return hours if hours >= 0 else None


def _parse_datetime(value: Any) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if hasattr(value, "year") and hasattr(value, "month"):  # datetime.date
        return datetime(value.year, value.month, value.day)
    text = _clean(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _score_header(headers: List[str], tokens: Tuple[str, ...]) -> int:
    joined = " | ".join(h.lower() for h in headers if h)
    return sum(1 for token in tokens if token in joined)


class RoadmapImportService:
    def __init__(self, db: Session):
        self.db = db

    # ============================================================== preview

    def build_preview(self, filename: str, content: bytes) -> RoadmapImportPreview:
        name = (filename or "").lower()
        default_title = re.sub(r"[_\-]+", " ", (filename or "Imported Roadmap").rsplit(".", 1)[0]).strip()

        if name.endswith((".xlsx", ".xlsm")):
            preview = self._preview_from_excel(content, default_title)
        elif name.endswith(".json"):
            preview = self._preview_from_json(content, default_title)
        elif name.endswith((".md", ".markdown")):
            preview = self._preview_from_markdown(content, default_title)
        elif name.endswith(".csv"):
            preview = self._preview_from_csv(content, default_title)
        else:
            raise ImportValidationException(
                "Unsupported roadmap file type. Use .xlsx, .json, .md, or .csv."
            )

        preview.source_filename = filename
        if not preview.topics:
            preview.warnings.append(
                "No topics were found in this file. Check that it has a table with "
                "Phase and Topic columns."
            )
        return preview

    # -------------------------------------------------------------- excel

    def _preview_from_excel(self, content: bytes, default_title: str) -> RoadmapImportPreview:
        # data_only=True yields the cached results of formulas. A progress
        # sheet built from cross-sheet references (as the reference workbook's
        # is) reads as None here when the file has never been opened and saved
        # by Excel -- handled explicitly below rather than importing blanks.
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        except Exception as exc:
            logger.error(f"Roadmap import: could not open workbook: {exc}")
            raise ImportValidationException(f"Could not read the Excel file: {exc}")

        warnings: List[str] = []
        ignored_sheets: List[str] = []
        syllabus_rows: List[RoadmapImportTopic] = []
        tracker_records: List[Dict[str, Any]] = []
        tracker_sheets: List[str] = []
        resources: List[RoadmapImportResource] = []

        for worksheet in workbook.worksheets:
            header_row, headers, kind = self._classify_sheet(worksheet)

            if kind == "syllabus":
                syllabus_rows.extend(self._read_syllabus(worksheet, header_row, headers, warnings))
            elif kind == "tracker":
                tracker_sheets.append(worksheet.title)
                tracker_records.extend(self._read_tracker(worksheet, header_row, headers))
            elif kind == "resource":
                resource = self._read_resource(worksheet, header_row, headers)
                if resource:
                    resources.append(resource)
            else:
                ignored_sheets.append(worksheet.title)

        if not syllabus_rows and tracker_records:
            # A tracker alone still describes topics and phases; use it as the
            # curriculum rather than refusing the file.
            syllabus_rows = [
                RoadmapImportTopic(
                    title=rec["title"],
                    phase_name=rec.get("phase_name") or DEFAULT_PHASE_NAME,
                    status=rec["status"],
                    progress_percentage=rec["progress_percentage"],
                    started_at=rec.get("started_at"),
                    completed_at=rec.get("completed_at"),
                    evidence_notes=rec.get("evidence_notes"),
                )
                for rec in tracker_records if rec.get("title")
            ]
        elif tracker_records:
            self._merge_tracker(syllabus_rows, tracker_records, warnings)

        if tracker_sheets and not tracker_records:
            # The sheet was recognised but every row read as empty. That is the
            # signature of formulas with no cached result -- openpyxl writes
            # none, so a workbook generated by a script (rather than saved by
            # Excel) looks exactly like this. Say so instead of quietly
            # importing a roadmap with all progress reset to zero.
            warnings.append(
                f"Progress sheet '{tracker_sheets[0]}' was found but contained no readable "
                "values, so only the syllabus was imported. If it uses formulas, open and "
                "re-save the file in Excel to cache their results, then import again."
            )

        title = self._workbook_title(workbook, default_title)
        return RoadmapImportPreview(
            title=title,
            phases=self._ordered_phase_names(syllabus_rows),
            topics=syllabus_rows,
            resources=resources,
            warnings=warnings,
            ignored_sheets=ignored_sheets,
        )

    @staticmethod
    def _workbook_title(workbook, default_title: str) -> str:
        return default_title or (workbook.worksheets[0].title if workbook.worksheets else "Imported Roadmap")

    def _classify_sheet(self, worksheet) -> Tuple[int, List[str], str]:
        """
        Find the header row and decide what kind of table this sheet holds.

        The header is not assumed to be row 1: the reference workbook's tracker
        sheet opens with a three-row summary block and only reaches its real
        header on row 5. The first ~10 rows are scored and the best match wins.
        """
        best: Tuple[int, List[str], str, int] = (1, [], "ignored", 0)

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            headers = [_clean(cell) or "" for cell in row]
            if not any(headers):
                continue

            syllabus_score = _score_header(headers, SYLLABUS_TOKENS)
            tracker_score = _score_header(headers, TRACKER_TOKENS)

            # A tracker sheet also says "Topic" and "Phase", so compare the two
            # signatures against each other instead of testing them in isolation.
            if tracker_score >= 3 and tracker_score > syllabus_score:
                kind, score = "tracker", tracker_score
            elif syllabus_score >= 3:
                kind, score = "syllabus", syllabus_score
            else:
                continue

            if score > best[3]:
                best = (row_idx, headers, kind, score)

        if best[2] != "ignored":
            return best[0], best[1], best[2]

        # Neither signature matched. A narrow two-to-four column table is a
        # reference/cheat-sheet -- keeping it means importing the reference
        # workbook doesn't silently discard half its sheets.
        first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first_row:
            headers = [_clean(cell) or "" for cell in first_row]
            populated = [h for h in headers if h]
            if 2 <= len(populated) <= 4 and worksheet.max_row > 1:
                return 1, headers, "resource"

        return 1, [], "ignored"

    @staticmethod
    def _column_map(headers: List[str]) -> Dict[str, int]:
        mapping: Dict[str, int] = {}
        for index, header in enumerate(headers):
            lowered = (header or "").lower()
            if not lowered:
                continue
            if "phase" in lowered or "module" in lowered or "section" in lowered:
                mapping.setdefault("phase", index)
            elif "objective" in lowered or "goal" in lowered:
                mapping.setdefault("objective", index)
            elif "success" in lowered or "criteria" in lowered or "outcome" in lowered:
                mapping.setdefault("success", index)
            elif "hour" in lowered or "effort" in lowered or "estimate" in lowered:
                mapping.setdefault("hours", index)
            elif "status" in lowered:
                mapping.setdefault("status", index)
            elif "progress" in lowered or "%" in lowered:
                mapping.setdefault("progress", index)
            elif "completion" in lowered or "completed" in lowered or "finish" in lowered:
                mapping.setdefault("completed_at", index)
            elif "start" in lowered:
                mapping.setdefault("started_at", index)
            elif "evidence" in lowered or "note" in lowered or "lab" in lowered:
                mapping.setdefault("evidence", index)
            elif "topic" in lowered or "title" in lowered or "skill" in lowered:
                mapping.setdefault("topic", index)
        return mapping

    def _is_summary_row(self, title: str, row: tuple, columns: Dict[str, int]) -> bool:
        """
        Distinguish a trailing totals line from a topic that merely starts with
        one of those words.

        Deliberately conservative: the title must look like a summary label
        *and* the row must be bare of the context a real topic carries. A topic
        called "Total Order Guarantees" sitting in a phase with an objective is
        kept; "TOTAL ESTIMATED HOURS" alone in an otherwise empty row is not.
        """
        if not SUMMARY_ROW_PATTERN.match(title):
            return False
        for key in ("phase", "objective", "success"):
            index = columns.get(key)
            if index is not None and _clean(self._cell(row, index)):
                return False
        return True

    def _read_syllabus(
        self, worksheet, header_row: int, headers: List[str], warnings: List[str],
    ) -> List[RoadmapImportTopic]:
        columns = self._column_map(headers)
        if "topic" not in columns:
            return []

        topics: List[RoadmapImportTopic] = []
        skipped: List[str] = []

        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            title = _clean(self._cell(row, columns.get("topic")))
            if not title:
                continue
            if self._is_summary_row(title, row, columns):
                skipped.append(title)
                continue
            topics.append(RoadmapImportTopic(
                title=title,
                phase_name=_clean(self._cell(row, columns.get("phase"))) or DEFAULT_PHASE_NAME,
                learning_objective=_clean(self._cell(row, columns.get("objective"))),
                success_criteria=_clean(self._cell(row, columns.get("success"))),
                estimated_hours=_parse_hours(self._cell(row, columns.get("hours"))),
            ))

        if skipped:
            warnings.append(
                f"Ignored {len(skipped)} summary row(s) in '{worksheet.title}': {', '.join(skipped[:3])}"
            )
        return topics

    def _read_tracker(self, worksheet, header_row: int, headers: List[str]) -> List[Dict[str, Any]]:
        columns = self._column_map(headers)
        records: List[Dict[str, Any]] = []
        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            title = _clean(self._cell(row, columns.get("topic")))
            status_raw = self._cell(row, columns.get("status"))
            if not title and status_raw is None:
                continue
            records.append({
                "title": title,
                "phase_name": _clean(self._cell(row, columns.get("phase"))),
                "status": _parse_status(status_raw),
                "progress_percentage": _parse_progress(self._cell(row, columns.get("progress"))),
                "started_at": _parse_datetime(self._cell(row, columns.get("started_at"))),
                "completed_at": _parse_datetime(self._cell(row, columns.get("completed_at"))),
                "evidence_notes": _clean(self._cell(row, columns.get("evidence"))),
            })
        return records

    @staticmethod
    def _cell(row: tuple, index: Optional[int]) -> Any:
        if index is None or index >= len(row):
            return None
        return row[index]

    @staticmethod
    def _merge_tracker(
        topics: List[RoadmapImportTopic],
        records: List[Dict[str, Any]],
        warnings: List[str],
    ) -> bool:
        """
        Fold recorded progress onto the syllabus, matching on topic title.

        Title matching rather than row position, so a tracker that has been
        re-sorted or partially filled still lands on the right rows. Returns
        whether anything actually matched.
        """
        by_title = {t.title.strip().lower(): t for t in topics}
        matched = 0
        unmatched: List[str] = []

        for record in records:
            title = (record.get("title") or "").strip().lower()
            target = by_title.get(title)
            if not target:
                if record.get("title"):
                    unmatched.append(record["title"])
                continue
            target.status = record["status"]
            target.progress_percentage = record["progress_percentage"]
            target.started_at = record.get("started_at")
            target.completed_at = record.get("completed_at")
            target.evidence_notes = record.get("evidence_notes")
            matched += 1

        if unmatched:
            preview = ", ".join(unmatched[:5])
            more = f" (+{len(unmatched) - 5} more)" if len(unmatched) > 5 else ""
            warnings.append(
                f"{len(unmatched)} progress row(s) did not match any syllabus topic and were "
                f"ignored: {preview}{more}"
            )
        return matched > 0

    def _read_resource(self, worksheet, header_row: int, headers: List[str]) -> Optional[RoadmapImportResource]:
        populated = [h for h in headers if h]
        if not populated:
            return None
        width = len(populated)
        rows: List[List[str]] = []
        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            values = [_clean(cell) or "" for cell in row[:width]]
            if any(values):
                rows.append(values)
        if not rows:
            return None
        return RoadmapImportResource(title=worksheet.title, columns=populated, rows=rows)

    # --------------------------------------------------------------- json

    def _preview_from_json(self, content: bytes, default_title: str) -> RoadmapImportPreview:
        try:
            data = json.loads(content.decode("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError) as exc:
            raise ImportValidationException(f"Invalid JSON roadmap file: {exc}")

        if not isinstance(data, dict):
            raise ImportValidationException("A JSON roadmap must be an object with a 'topics' or 'phases' key.")

        topics: List[RoadmapImportTopic] = []

        # Nested {phases: [{name, topics: []}]} and flat {topics: [{phase}]}
        # are both accepted -- the nested form preserves phase order explicitly.
        for phase in data.get("phases") or []:
            phase_name = _clean(phase.get("name")) or DEFAULT_PHASE_NAME
            for topic in phase.get("topics") or []:
                topics.append(self._json_topic(topic, phase_name))

        for topic in data.get("topics") or []:
            phase_name = _clean(topic.get("phase") or topic.get("phase_name")) or DEFAULT_PHASE_NAME
            topics.append(self._json_topic(topic, phase_name))

        resources = [
            RoadmapImportResource(
                title=_clean(r.get("title")) or "Reference",
                columns=[str(c) for c in (r.get("columns") or [])],
                rows=[[_clean(c) or "" for c in row] for row in (r.get("rows") or [])],
            )
            for r in data.get("resources") or []
        ]

        return RoadmapImportPreview(
            title=_clean(data.get("title")) or default_title,
            description=_clean(data.get("description")),
            phases=self._ordered_phase_names(topics),
            topics=topics,
            resources=resources,
        )

    @staticmethod
    def _json_topic(payload: dict, phase_name: str) -> RoadmapImportTopic:
        return RoadmapImportTopic(
            title=_clean(payload.get("title") or payload.get("topic")) or "Untitled Topic",
            phase_name=phase_name,
            learning_objective=_clean(payload.get("learning_objective") or payload.get("objective")),
            success_criteria=_clean(payload.get("success_criteria") or payload.get("success")),
            estimated_hours=_parse_hours(payload.get("estimated_hours") or payload.get("hours")),
            status=_parse_status(payload.get("status")),
            progress_percentage=_parse_progress(payload.get("progress_percentage") or payload.get("progress")),
            evidence_notes=_clean(payload.get("evidence_notes")),
        )

    # ----------------------------------------------------------- markdown

    def _preview_from_markdown(self, content: bytes, default_title: str) -> RoadmapImportPreview:
        text = content.decode("utf-8", errors="replace")
        title = default_title
        current_phase = DEFAULT_PHASE_NAME
        topics: List[RoadmapImportTopic] = []

        for raw_line in text.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            if line.startswith("# ") and not line.startswith("## "):
                title = line[2:].strip() or title
            elif line.startswith("## ") and not line.startswith("### "):
                current_phase = line[3:].strip() or DEFAULT_PHASE_NAME
            elif line.startswith("### "):
                topics.append(self._markdown_topic(line[4:].strip(), current_phase))
            elif line.startswith(("- ", "* ")):
                topics.append(self._markdown_topic(line[2:].strip(), current_phase))

        return RoadmapImportPreview(
            title=title,
            phases=self._ordered_phase_names(topics),
            topics=topics,
        )

    @staticmethod
    def _markdown_topic(line: str, phase_name: str) -> RoadmapImportTopic:
        # "Consumer Internals (4h)" / "Consumer Internals - 4 hours"
        hours = None
        match = re.search(r"[\(\-–—]\s*(\d+(?:\.\d+)?)\s*(?:h|hr|hrs|hour|hours)\b", line, re.IGNORECASE)
        if match:
            hours = float(match.group(1))
            line = line[:match.start()].strip()
        checked = None
        if line.lower().startswith(("[x]", "[X]")):
            checked, line = RoadmapTopicStatus.COMPLETED, line[3:].strip()
        elif line.startswith("[ ]"):
            checked, line = RoadmapTopicStatus.NOT_STARTED, line[3:].strip()
        return RoadmapImportTopic(
            title=line or "Untitled Topic",
            phase_name=phase_name,
            estimated_hours=hours,
            status=checked or RoadmapTopicStatus.NOT_STARTED,
            progress_percentage=100 if checked == RoadmapTopicStatus.COMPLETED else 0,
        )

    # ---------------------------------------------------------------- csv

    def _preview_from_csv(self, content: bytes, default_title: str) -> RoadmapImportPreview:
        try:
            frame = pd.read_csv(io.BytesIO(content))
        except Exception as exc:
            raise ImportValidationException(f"Could not read the CSV file: {exc}")

        headers = [str(c) for c in frame.columns]
        columns = self._column_map(headers)
        if "topic" not in columns:
            raise ImportValidationException(
                "The CSV needs a Topic (or Title) column to import as a roadmap."
            )

        topics: List[RoadmapImportTopic] = []
        warnings: List[str] = []
        skipped: List[str] = []
        for _, row in frame.iterrows():
            values = tuple(row.values)
            title = _clean(self._cell(values, columns.get("topic")))
            if not title:
                continue
            if self._is_summary_row(title, values, columns):
                skipped.append(title)
                continue
            values = list(values)
            topics.append(RoadmapImportTopic(
                title=title,
                phase_name=_clean(self._cell(tuple(values), columns.get("phase"))) or DEFAULT_PHASE_NAME,
                learning_objective=_clean(self._cell(tuple(values), columns.get("objective"))),
                success_criteria=_clean(self._cell(tuple(values), columns.get("success"))),
                estimated_hours=_parse_hours(self._cell(tuple(values), columns.get("hours"))),
                status=_parse_status(self._cell(tuple(values), columns.get("status"))),
                progress_percentage=_parse_progress(self._cell(tuple(values), columns.get("progress"))),
            ))

        if skipped:
            warnings.append(f"Ignored {len(skipped)} summary row(s): {', '.join(skipped[:3])}")

        return RoadmapImportPreview(
            title=default_title,
            phases=self._ordered_phase_names(topics),
            topics=topics,
            warnings=warnings,
        )

    # ------------------------------------------------------------ helpers

    @staticmethod
    def _ordered_phase_names(topics: List[RoadmapImportTopic]) -> List[str]:
        """Unique phase names in first-appearance order -- that ordering is the
        curriculum's own sequence and must survive the round trip."""
        seen: List[str] = []
        for topic in topics:
            if topic.phase_name not in seen:
                seen.append(topic.phase_name)
        return seen

    # =============================================================== commit

    def commit(self, req: RoadmapImportConfirm) -> RoadmapImportResult:
        """
        Write the approved preview as one roadmap, in a single transaction --
        a failure part-way through rolls the whole thing back rather than
        leaving a half-built roadmap the user has to find and delete.
        """
        if not req.topics:
            raise ImportValidationException("Cannot import a roadmap with no topics.")

        try:
            roadmap = Roadmap(
                title=req.title.strip() or "Imported Roadmap",
                description=req.description,
                source_filename=req.source_filename,
                start_date=req.start_date,
                weekly_hours_budget=req.weekly_hours_budget,
            )
            self.db.add(roadmap)
            self.db.flush()

            phases: Dict[str, RoadmapPhase] = {}
            topic_counters: Dict[str, int] = {}

            for item in req.topics:
                phase_name = (item.phase_name or DEFAULT_PHASE_NAME).strip() or DEFAULT_PHASE_NAME

                phase = phases.get(phase_name)
                if phase is None:
                    phase = RoadmapPhase(
                        roadmap_id=roadmap.id,
                        name=phase_name,
                        order_index=len(phases),
                    )
                    self.db.add(phase)
                    self.db.flush()
                    phases[phase_name] = phase
                    topic_counters[phase_name] = 0

                topic = RoadmapTopic(
                    roadmap_id=roadmap.id,
                    phase_id=phase.id,
                    order_index=topic_counters[phase_name],
                    title=item.title.strip(),
                    learning_objective=item.learning_objective,
                    success_criteria=item.success_criteria,
                    estimated_hours=item.estimated_hours,
                    status=item.status,
                    progress_percentage=item.progress_percentage,
                    started_at=item.started_at,
                    completed_at=item.completed_at,
                    evidence_notes=item.evidence_notes,
                )
                # Coherence only. _stamp_transition_times is deliberately not
                # called: an imported "Completed" row with no date in the file
                # must not be stamped with today's date.
                RoadmapService._reconcile_topic_state(topic)
                self.db.add(topic)
                topic_counters[phase_name] += 1

            for index, resource in enumerate(req.resources):
                self.db.add(RoadmapResource(
                    roadmap_id=roadmap.id,
                    title=resource.title,
                    order_index=index,
                    columns=resource.columns,
                    rows=resource.rows,
                ))

            self.db.commit()
            self.db.refresh(roadmap)

            return RoadmapImportResult(
                roadmap_id=roadmap.id,
                title=roadmap.title,
                phase_count=len(phases),
                topic_count=len(req.topics),
                resource_count=len(req.resources),
            )
        except Exception as exc:
            self.db.rollback()
            logger.error(f"Roadmap import commit failed: {exc}")
            raise
