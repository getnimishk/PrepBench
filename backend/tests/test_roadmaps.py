"""
test_roadmaps.py

Covers the Learning Roadmap vertical slice: progress math, the
status/progress invariant, schedule derivation, and the generic importer.

Workbooks are built in memory with openpyxl rather than committed as
fixtures, so the tests describe the shapes they care about (a totals row, a
formula-only progress sheet, a missing hours column) instead of depending on
one particular file.
"""

import io
import json
import openpyxl
import pytest
from datetime import date, datetime, timedelta
from fastapi.testclient import TestClient

from app.main import app
from app.models.roadmap import Roadmap, RoadmapPhase, RoadmapTopic, RoadmapResource
from tests.conftest import TestingSessionLocal

client = TestClient(app)


@pytest.fixture
def roadmap_ids():
    """Deletes roadmaps created by a test. Children go via ON DELETE CASCADE."""
    created = []
    yield created
    db = TestingSessionLocal()
    try:
        if created:
            db.query(Roadmap).filter(Roadmap.id.in_(created)).delete(synchronize_session=False)
            db.commit()
    finally:
        db.close()


def _make_workbook(sheets: dict) -> bytes:
    """sheets: {sheet_name: [row, row, ...]} -- first row is the header."""
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets.items():
        worksheet = workbook.create_sheet(title=name)
        for row in rows:
            worksheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


SYLLABUS_HEADER = ["Phase", "#", "Topic", "Learning Objective", "Success Criteria", "Est. Hours"]


def _syllabus_rows():
    return [
        SYLLABUS_HEADER,
        ["1. Foundations", 1, "Event Streaming Basics", "Understand pub/sub.", "Explain it.", 3],
        ["1. Foundations", 2, "Record Structure", "Key/value/headers.", "Trace a record.", 2],
        ["2. Delivery", 3, "Offset Commits", "Manual vs auto.", "Commit correctly.", 4],
    ]


def _import(content: bytes, filename: str, roadmap_ids, **overrides):
    """validate -> confirm, returning the created roadmap's detail payload."""
    res = client.post("/api/v1/roadmaps/import/validate",
                      files={"file": (filename, content, "application/octet-stream")})
    assert res.status_code == 200, res.text
    preview = res.json()

    body = {
        "title": preview["title"],
        "source_filename": preview["source_filename"],
        "topics": preview["topics"],
        "resources": preview["resources"],
        **overrides,
    }
    confirm = client.post("/api/v1/roadmaps/import/confirm", json=body)
    assert confirm.status_code == 201, confirm.text
    roadmap_id = confirm.json()["roadmap_id"]
    roadmap_ids.append(roadmap_id)
    return preview, client.get(f"/api/v1/roadmaps/{roadmap_id}").json()


# ============================================================ import: excel

def test_excel_import_builds_phases_topics_and_preserves_order(roadmap_ids):
    content = _make_workbook({"Syllabus": _syllabus_rows()})
    preview, detail = _import(content, "kafka_roadmap.xlsx", roadmap_ids)

    assert preview["phases"] == ["1. Foundations", "2. Delivery"]
    assert len(preview["topics"]) == 3

    assert [p["name"] for p in detail["phases"]] == ["1. Foundations", "2. Delivery"]
    assert [t["title"] for t in detail["phases"][0]["topics"]] == [
        "Event Streaming Basics", "Record Structure",
    ]
    assert detail["phases"][0]["topics"][0]["estimated_hours"] == 3
    assert detail["phases"][0]["topics"][0]["learning_objective"] == "Understand pub/sub."


def test_excel_import_ignores_a_trailing_totals_row(roadmap_ids):
    """
    A totals line imported as a topic becomes a phantom 9-hour item in a
    phantom phase, which then corrupts every hours-derived figure.
    """
    rows = _syllabus_rows() + [[None, None, "TOTAL ESTIMATED HOURS", None, None, 9]]
    preview, detail = _import(_make_workbook({"Syllabus": rows}), "roadmap.xlsx", roadmap_ids)

    titles = [t["title"] for t in preview["topics"]]
    assert "TOTAL ESTIMATED HOURS" not in titles
    assert len(titles) == 3
    assert detail["progress"]["total_estimated_hours"] == 9  # 3 + 2 + 4, not 18
    assert any("summary row" in w for w in preview["warnings"])


def test_a_topic_merely_starting_with_total_is_not_treated_as_a_summary_row(roadmap_ids):
    rows = [
        SYLLABUS_HEADER,
        ["3. Ordering", 1, "Total Order Guarantees", "Understand global ordering.", "Explain it.", 5],
    ]
    preview, _ = _import(_make_workbook({"Syllabus": rows}), "roadmap.xlsx", roadmap_ids)
    assert [t["title"] for t in preview["topics"]] == ["Total Order Guarantees"]


def test_sheets_are_classified_by_header_shape_not_by_sheet_name(roadmap_ids):
    """
    The reference workbook names its sheets "Kafka Master Syllabus" and
    "Journey Progress Tracker". Matching on names like those would silently
    make the importer work for exactly one roadmap.
    """
    content = _make_workbook({
        "Some Arbitrary Name": _syllabus_rows(),
        "Totally Unrelated Title": [
            ["#", "Topic", "Phase", "Status", "Progress %", "Start Date", "Completion Date", "Evidence"],
            [1, "Event Streaming Basics", "1. Foundations", "Completed", 100, None, None, "Wrote a demo"],
            [2, "Record Structure", "1. Foundations", "In Progress", 40, None, None, None],
        ],
    })
    preview, detail = _import(content, "roadmap.xlsx", roadmap_ids)

    by_title = {t["title"]: t for t in preview["topics"]}
    assert by_title["Event Streaming Basics"]["status"] == "completed"
    assert by_title["Event Streaming Basics"]["evidence_notes"] == "Wrote a demo"
    assert by_title["Record Structure"]["status"] == "in_progress"
    assert by_title["Record Structure"]["progress_percentage"] == 40
    # The syllabus columns still came through, so the tracker merged onto the
    # curriculum rather than replacing it.
    assert by_title["Event Streaming Basics"]["estimated_hours"] == 3
    assert detail["progress"]["completed_count"] == 1


def test_importing_a_completed_row_without_a_date_does_not_stamp_today(roadmap_ids):
    """
    The spreadsheet says "Completed" but records no date. Stamping today would
    assert you finished it on import day and would draw the Gantt's "actual"
    bar in the wrong place.
    """
    content = _make_workbook({
        "Syllabus": _syllabus_rows(),
        "Progress": [
            ["#", "Topic", "Status", "Progress %", "Start Date", "Completion Date"],
            [1, "Event Streaming Basics", "Completed", 100, None, None],
        ],
    })
    _, detail = _import(content, "roadmap.xlsx", roadmap_ids)

    topic = next(t for p in detail["phases"] for t in p["topics"] if t["title"] == "Event Streaming Basics")
    assert topic["status"] == "completed"
    assert topic["progress_percentage"] == 100
    assert topic["completed_at"] is None
    assert topic["started_at"] is None


def test_narrow_reference_sheets_are_preserved_as_resources(roadmap_ids):
    """Otherwise importing the reference workbook silently discards half of it."""
    content = _make_workbook({
        "Syllabus": _syllabus_rows(),
        "CLI Command Reference": [
            ["Category", "Linux", "Windows"],
            ["Create Topic", "kafka-topics.sh --create", "kafka-topics.bat --create"],
        ],
        "Mental Model": [
            ["Concept", "Mental Model"],
            ["Topic", "Database table for events"],
        ],
    })
    preview, detail = _import(content, "roadmap.xlsx", roadmap_ids)

    titles = {r["title"] for r in preview["resources"]}
    assert titles == {"CLI Command Reference", "Mental Model"}
    cli = next(r for r in detail["resources"] if r["title"] == "CLI Command Reference")
    assert cli["columns"] == ["Category", "Linux", "Windows"]
    assert cli["rows"] == [["Create Topic", "kafka-topics.sh --create", "kafka-topics.bat --create"]]


def test_progress_sheet_with_uncached_formulas_warns_instead_of_importing_blanks():
    """
    openpyxl writes formulas with no cached result, which is exactly what a
    programmatically-generated workbook looks like. Reading it with
    data_only=True yields None for every formula cell, so the merge must be
    skipped with an explanation rather than silently wiping progress.
    """
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    syllabus = workbook.create_sheet("Syllabus")
    for row in _syllabus_rows():
        syllabus.append(row)

    tracker = workbook.create_sheet("Tracker")
    tracker.append(["#", "Topic", "Status", "Progress %", "Start Date", "Completion Date"])
    for index in range(2, 5):
        tracker.append([index - 1, f"=Syllabus!C{index}", "=Syllabus!A2", None, None, None])

    buffer = io.BytesIO()
    workbook.save(buffer)

    res = client.post("/api/v1/roadmaps/import/validate",
                      files={"file": ("roadmap.xlsx", buffer.getvalue(), "application/octet-stream")})
    assert res.status_code == 200
    preview = res.json()

    assert len(preview["topics"]) == 3
    assert all(t["status"] == "not_started" for t in preview["topics"])
    assert any("re-save" in w.lower() for w in preview["warnings"])


def test_unsupported_file_type_is_rejected():
    res = client.post("/api/v1/roadmaps/import/validate",
                      files={"file": ("roadmap.pdf", b"%PDF-1.4", "application/pdf")})
    assert res.status_code == 422


# ===================================================== import: other formats

def test_json_import_supports_nested_phases(roadmap_ids):
    payload = {
        "title": "Rust Mastery",
        "description": "From zero to async",
        "phases": [
            {"name": "Basics", "topics": [
                {"title": "Ownership", "hours": 4, "objective": "Understand moves."},
                {"title": "Borrowing", "hours": 3},
            ]},
            {"name": "Async", "topics": [{"title": "Futures", "hours": 6, "status": "in_progress", "progress": 25}]},
        ],
    }
    _, detail = _import(json.dumps(payload).encode(), "rust.json", roadmap_ids)

    assert detail["title"] == "Rust Mastery"
    assert [p["name"] for p in detail["phases"]] == ["Basics", "Async"]
    assert detail["progress"]["total_estimated_hours"] == 13
    futures = detail["phases"][1]["topics"][0]
    assert futures["status"] == "in_progress"
    assert futures["progress_percentage"] == 25


def test_markdown_import_reads_headings_checkboxes_and_hours(roadmap_ids):
    markdown = (
        "# Go Roadmap\n"
        "## Fundamentals\n"
        "- [x] Goroutines (3h)\n"
        "- [ ] Channels (2h)\n"
        "## Tooling\n"
        "### Modules\n"
    )
    _, detail = _import(markdown.encode(), "go.md", roadmap_ids)

    assert detail["title"] == "Go Roadmap"
    assert [p["name"] for p in detail["phases"]] == ["Fundamentals", "Tooling"]
    goroutines, channels = detail["phases"][0]["topics"]
    assert (goroutines["title"], goroutines["status"], goroutines["estimated_hours"]) == ("Goroutines", "completed", 3)
    assert (channels["title"], channels["status"], channels["estimated_hours"]) == ("Channels", "not_started", 2)


def test_csv_import_without_a_topic_column_is_rejected():
    res = client.post("/api/v1/roadmaps/import/validate",
                      files={"file": ("bad.csv", b"Alpha,Beta\n1,2\n", "text/csv")})
    assert res.status_code == 422


def test_importing_a_roadmap_with_no_topics_is_rejected():
    res = client.post("/api/v1/roadmaps/import/confirm",
                      json={"title": "Empty", "topics": [], "resources": []})
    assert res.status_code == 422


# ============================================================ progress math

def _new_roadmap(roadmap_ids, **kwargs) -> int:
    res = client.post("/api/v1/roadmaps", json={"title": "Manual Roadmap", **kwargs})
    assert res.status_code == 201
    roadmap_id = res.json()["id"]
    roadmap_ids.append(roadmap_id)
    return roadmap_id


def _add_topics(roadmap_id: int, hours_list):
    phase = client.post(f"/api/v1/roadmaps/{roadmap_id}/phases", json={"name": "Phase 1"})
    phase_id = phase.json()["id"]
    ids = []
    for index, hours in enumerate(hours_list):
        res = client.post(f"/api/v1/roadmaps/{roadmap_id}/topics",
                          json={"phase_id": phase_id, "title": f"Topic {index}", "estimated_hours": hours})
        assert res.status_code == 201
        ids.append(res.json()["id"])
    return phase_id, ids


def test_empty_roadmap_reports_null_progress_not_zero(roadmap_ids):
    """
    A roadmap with no topics is not "0% complete" -- that asserts progress it
    has no basis to measure, the same error as showing a fabricated score for
    an ungraded attempt.
    """
    roadmap_id = _new_roadmap(roadmap_ids)
    progress = client.get(f"/api/v1/roadmaps/{roadmap_id}").json()["progress"]

    assert progress["total_topics"] == 0
    assert progress["completion_percentage"] is None
    assert progress["hours_percentage"] is None
    assert progress["total_estimated_hours"] is None


def test_hours_percentage_is_null_unless_every_topic_has_an_estimate(roadmap_ids):
    """All-or-nothing: an hours figure from only the topics that happen to
    carry estimates is a confidently wrong number."""
    roadmap_id = _new_roadmap(roadmap_ids)
    _add_topics(roadmap_id, [3, None, 5])

    progress = client.get(f"/api/v1/roadmaps/{roadmap_id}").json()["progress"]
    assert progress["completion_percentage"] == 0.0   # countable, just none done
    assert progress["hours_percentage"] is None
    assert progress["total_estimated_hours"] is None


def test_skipped_topics_leave_the_denominator_entirely(roadmap_ids):
    roadmap_id = _new_roadmap(roadmap_ids)
    _, topic_ids = _add_topics(roadmap_id, [2, 2, 2, 2])

    client.patch(f"/api/v1/roadmaps/{roadmap_id}/topics/{topic_ids[0]}", json={"status": "completed"})
    client.patch(f"/api/v1/roadmaps/{roadmap_id}/topics/{topic_ids[1]}", json={"status": "skipped"})

    progress = client.get(f"/api/v1/roadmaps/{roadmap_id}").json()["progress"]
    # 1 of 3 countable, not 1 of 4 -- material you already know must not pin
    # the roadmap below 100% forever.
    assert progress["completion_percentage"] == pytest.approx(33.3)
    assert progress["skipped_count"] == 1
    assert progress["total_estimated_hours"] == 6
    assert progress["hours_percentage"] == pytest.approx(33.3)


def test_a_fully_skipped_roadmap_reports_null_rather_than_zero(roadmap_ids):
    roadmap_id = _new_roadmap(roadmap_ids)
    _, topic_ids = _add_topics(roadmap_id, [1, 1])
    for topic_id in topic_ids:
        client.patch(f"/api/v1/roadmaps/{roadmap_id}/topics/{topic_id}", json={"status": "skipped"})

    progress = client.get(f"/api/v1/roadmaps/{roadmap_id}").json()["progress"]
    assert progress["completion_percentage"] is None
    assert progress["hours_percentage"] is None


# ====================================================== status/progress rules

def test_completing_a_topic_forces_full_progress_and_stamps_dates(roadmap_ids):
    roadmap_id = _new_roadmap(roadmap_ids)
    _, topic_ids = _add_topics(roadmap_id, [3])

    body = client.patch(f"/api/v1/roadmaps/{roadmap_id}/topics/{topic_ids[0]}",
                        json={"status": "completed"}).json()
    assert body["progress_percentage"] == 100
    assert body["completed_at"] is not None
    assert body["started_at"] is not None


def test_setting_progress_alone_infers_the_status(roadmap_ids):
    roadmap_id = _new_roadmap(roadmap_ids)
    _, topic_ids = _add_topics(roadmap_id, [3])
    topic_id = topic_ids[0]

    partial = client.patch(f"/api/v1/roadmaps/{roadmap_id}/topics/{topic_id}",
                           json={"progress_percentage": 40}).json()
    assert partial["status"] == "in_progress"
    assert partial["started_at"] is not None
    assert partial["completed_at"] is None

    done = client.patch(f"/api/v1/roadmaps/{roadmap_id}/topics/{topic_id}",
                        json={"progress_percentage": 100}).json()
    assert done["status"] == "completed"
    assert done["completed_at"] is not None


def test_in_progress_at_100_percent_is_clamped_rather_than_silently_completed(roadmap_ids):
    """The caller asked for in_progress explicitly; honour that but refuse the
    contradictory number."""
    roadmap_id = _new_roadmap(roadmap_ids)
    _, topic_ids = _add_topics(roadmap_id, [3])

    body = client.patch(f"/api/v1/roadmaps/{roadmap_id}/topics/{topic_ids[0]}",
                        json={"status": "in_progress", "progress_percentage": 100}).json()
    assert body["status"] == "in_progress"
    assert body["progress_percentage"] == 99


def test_reverting_to_not_started_clears_progress_and_dates(roadmap_ids):
    roadmap_id = _new_roadmap(roadmap_ids)
    _, topic_ids = _add_topics(roadmap_id, [3])
    topic_id = topic_ids[0]

    client.patch(f"/api/v1/roadmaps/{roadmap_id}/topics/{topic_id}", json={"status": "completed"})
    body = client.patch(f"/api/v1/roadmaps/{roadmap_id}/topics/{topic_id}",
                        json={"status": "not_started"}).json()

    assert body["progress_percentage"] == 0
    assert body["started_at"] is None
    assert body["completed_at"] is None


def test_patching_status_leaves_other_fields_untouched(roadmap_ids):
    """exclude_unset: omitting evidence_notes must not clear it."""
    roadmap_id = _new_roadmap(roadmap_ids)
    _, topic_ids = _add_topics(roadmap_id, [3])
    topic_id = topic_ids[0]

    client.patch(f"/api/v1/roadmaps/{roadmap_id}/topics/{topic_id}",
                 json={"evidence_notes": "Built a demo cluster"})
    body = client.patch(f"/api/v1/roadmaps/{roadmap_id}/topics/{topic_id}",
                        json={"status": "completed"}).json()

    assert body["evidence_notes"] == "Built a demo cluster"


# ================================================================= schedule

def test_schedule_reports_why_it_is_unavailable_rather_than_inventing_one(roadmap_ids):
    roadmap_id = _new_roadmap(roadmap_ids)

    assert client.get(f"/api/v1/roadmaps/{roadmap_id}/schedule").json()["reason"] == "no_topics"

    _add_topics(roadmap_id, [4, 4])
    body = client.get(f"/api/v1/roadmaps/{roadmap_id}/schedule").json()
    assert body["schedule_available"] is False
    assert body["reason"] == "no_start_date"

    client.put(f"/api/v1/roadmaps/{roadmap_id}", json={"start_date": "2026-01-01"})
    body = client.get(f"/api/v1/roadmaps/{roadmap_id}/schedule").json()
    assert body["schedule_available"] is False
    assert body["reason"] == "no_weekly_budget"


def test_schedule_without_any_hours_reports_no_time_estimates(roadmap_ids):
    roadmap_id = _new_roadmap(roadmap_ids, start_date="2026-01-01", weekly_hours_budget=10)
    _add_topics(roadmap_id, [None, None])

    body = client.get(f"/api/v1/roadmaps/{roadmap_id}/schedule").json()
    assert body["schedule_available"] is False
    assert body["reason"] == "no_time_estimates"


def test_schedule_projects_sequential_bars_from_the_weekly_budget(roadmap_ids):
    roadmap_id = _new_roadmap(roadmap_ids, start_date="2026-01-01", weekly_hours_budget=7)
    _add_topics(roadmap_id, [7, 7])

    body = client.get(f"/api/v1/roadmaps/{roadmap_id}/schedule").json()
    assert body["schedule_available"] is True

    first, second = body["items"]
    assert first["schedule_status"] == "projected"
    # 7h at 1h/day = 7 days, and the next bar starts the day after.
    assert (date.fromisoformat(first["end"]) - date.fromisoformat(first["start"])).days == 6
    assert date.fromisoformat(second["start"]) == date.fromisoformat(first["end"]) + timedelta(days=1)
    assert body["projected_end_date"] == second["end"]


def test_a_topic_without_hours_is_marked_unschedulable_not_given_a_default(roadmap_ids):
    """Defaulting a missing estimate to an hour would silently invent a
    delivery date."""
    roadmap_id = _new_roadmap(roadmap_ids, start_date="2026-01-01", weekly_hours_budget=7)
    _add_topics(roadmap_id, [7, None])

    body = client.get(f"/api/v1/roadmaps/{roadmap_id}/schedule").json()
    assert body["schedule_available"] is True
    assert body["unschedulable_topic_count"] == 1

    unschedulable = body["items"][1]
    assert unschedulable["schedule_status"] == "unschedulable"
    assert unschedulable["start"] is None and unschedulable["end"] is None


def test_completed_topics_use_actual_dates_and_consume_no_future_budget(roadmap_ids):
    roadmap_id = _new_roadmap(roadmap_ids, start_date="2026-01-01", weekly_hours_budget=7)
    _, topic_ids = _add_topics(roadmap_id, [7, 7])
    client.patch(f"/api/v1/roadmaps/{roadmap_id}/topics/{topic_ids[0]}", json={"status": "completed"})

    body = client.get(f"/api/v1/roadmaps/{roadmap_id}/schedule").json()
    done, upcoming = body["items"]

    assert done["schedule_status"] == "actual"
    assert done["start"] is not None
    # The remaining topic starts from today's forecast, not queued behind the
    # finished one's original slot.
    assert upcoming["schedule_status"] == "projected"
    assert date.fromisoformat(upcoming["start"]) >= date.today()


def test_skipped_topics_get_no_bar(roadmap_ids):
    roadmap_id = _new_roadmap(roadmap_ids, start_date="2026-01-01", weekly_hours_budget=7)
    _, topic_ids = _add_topics(roadmap_id, [7, 7])
    client.patch(f"/api/v1/roadmaps/{roadmap_id}/topics/{topic_ids[0]}", json={"status": "skipped"})

    body = client.get(f"/api/v1/roadmaps/{roadmap_id}/schedule").json()
    assert body["items"][0]["schedule_status"] == "skipped"
    assert body["items"][0]["start"] is None


# ==================================================================== CRUD

def test_list_reports_progress_per_roadmap_and_hides_archived(roadmap_ids):
    visible = _new_roadmap(roadmap_ids)
    archived = _new_roadmap(roadmap_ids)
    _add_topics(visible, [1, 1])
    client.put(f"/api/v1/roadmaps/{archived}", json={"is_archived": True})

    listed = client.get("/api/v1/roadmaps").json()
    ids = [r["id"] for r in listed]
    assert visible in ids and archived not in ids
    assert next(r for r in listed if r["id"] == visible)["progress"]["total_topics"] == 2

    with_archived = client.get("/api/v1/roadmaps", params={"include_archived": True}).json()
    assert archived in [r["id"] for r in with_archived]


def test_unknown_roadmap_returns_404():
    assert client.get("/api/v1/roadmaps/999999").status_code == 404
    assert client.get("/api/v1/roadmaps/999999/schedule").status_code == 404


def test_a_topic_cannot_be_moved_under_another_roadmaps_phase(roadmap_ids):
    """Otherwise roadmap_id and phase_id would point at different roadmaps."""
    first = _new_roadmap(roadmap_ids)
    second = _new_roadmap(roadmap_ids)
    _, topic_ids = _add_topics(first, [1])
    foreign_phase = client.post(f"/api/v1/roadmaps/{second}/phases", json={"name": "Other"}).json()["id"]

    res = client.patch(f"/api/v1/roadmaps/{first}/topics/{topic_ids[0]}", json={"phase_id": foreign_phase})
    assert res.status_code == 404


def test_deleting_a_roadmap_cascades_to_phases_topics_and_resources(roadmap_ids):
    content = _make_workbook({
        "Syllabus": _syllabus_rows(),
        "Reference": [["Concept", "Meaning"], ["Topic", "A log"]],
    })
    _, detail = _import(content, "roadmap.xlsx", roadmap_ids)
    roadmap_id = detail["id"]

    assert client.delete(f"/api/v1/roadmaps/{roadmap_id}").status_code == 204

    db = TestingSessionLocal()
    try:
        assert db.query(RoadmapPhase).filter(RoadmapPhase.roadmap_id == roadmap_id).count() == 0
        assert db.query(RoadmapTopic).filter(RoadmapTopic.roadmap_id == roadmap_id).count() == 0
        assert db.query(RoadmapResource).filter(RoadmapResource.roadmap_id == roadmap_id).count() == 0
    finally:
        db.close()
