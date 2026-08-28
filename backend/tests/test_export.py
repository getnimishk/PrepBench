# PrepBench - Copyright (c) 2026 Nimish Kanungo
# Licensed under the PolyForm Noncommercial License 1.0.0 (see LICENSE).
# Commercial use requires a separate licence from the copyright holder.

"""
test_export.py

Coverage for the PDF/Excel exam report export endpoints (app/api/v1/export.py,
app/services/export_service.py, app/utils/pdf_generator.py,
app/utils/excel_generator.py) -- previously untested.
"""

import io
import uuid
import openpyxl
import pytest
from fastapi.testclient import TestClient
from app.main import app
from app.models.exam_session import ExamSession
from app.models.exam_answer import ExamAnswer
from app.models.question import Question
from app.models.option import QuestionOption
from app.models.spaced_repetition import SpacedRepetition
from tests.conftest import TestingSessionLocal

client = TestClient(app)


@pytest.fixture
def cleanup_registry():
    """
    Tracks exam sessions/questions created during a test so they can be
    removed afterward. conftest.py's test database is created once per
    pytest session with no per-test rollback, so rows left behind here would
    otherwise persist for the rest of the run and could corrupt count-based
    assertions in other test modules depending on collection order (e.g.
    test_e2e_full_suite.py's `total == 0` check after clearing all questions).

    Children are deleted explicitly rather than relying on each model's
    `ondelete="CASCADE"` -- that only fires when SQLite's `PRAGMA
    foreign_keys` is ON, which is set on app.core.database's production
    `engine` via a connect-event listener that conftest.py's separate
    `test_engine` never receives. Without this, a bulk delete of just the
    Question row leaves its question_options/spaced_repetition rows
    orphaned; SQLite then reuses the freed question id on the next insert
    (ROWID reuse once a table is empty), silently attaching the new
    question to the old orphaned option rows too.
    """
    created = {"session_ids": [], "question_ids": []}
    yield created
    db = TestingSessionLocal()
    try:
        if created["session_ids"]:
            db.query(ExamAnswer).filter(ExamAnswer.session_id.in_(created["session_ids"])).delete(synchronize_session=False)
            db.query(ExamSession).filter(ExamSession.id.in_(created["session_ids"])).delete(synchronize_session=False)
        if created["question_ids"]:
            db.query(ExamAnswer).filter(ExamAnswer.question_id.in_(created["question_ids"])).delete(synchronize_session=False)
            db.query(SpacedRepetition).filter(SpacedRepetition.question_id.in_(created["question_ids"])).delete(synchronize_session=False)
            db.query(QuestionOption).filter(QuestionOption.question_id.in_(created["question_ids"])).delete(synchronize_session=False)
            db.query(Question).filter(Question.id.in_(created["question_ids"])).delete(synchronize_session=False)
        db.commit()
    finally:
        db.close()


def _create_completed_exam_session(registry):
    """
    Creates one dedicated question under a unique certification tag (so this
    exam session's question filter can't accidentally pick up questions
    created by other tests sharing the same persistent test database),
    answers it correctly, and finishes the exam. Returns the completed
    session's id together with its raw JSON for assertions. Registers both
    the question and the session with `registry` so `cleanup_registry`
    removes them after the test.
    """
    # A pure hex string (no shared static word) -- create_exam's certification
    # filter tokenizes on '-'/whitespace and does substring ilike matching, so
    # a shared prefix like "ExportTestCert-<hex>" across multiple tests would
    # cross-match every other test's questions via the "ExportTestCert" token.
    unique_cert = uuid.uuid4().hex
    q_payload = {
        "text": f"Export test question {uuid.uuid4().hex}",
        "question_type": "single_choice",
        "difficulty": "medium",
        "domain": "Export Domain",
        "topic": "Export Topic",
        "certification": unique_cert,
        "options": [
            {"option_text": "Correct Option", "is_correct": True, "order_index": 0},
            {"option_text": "Wrong Option", "is_correct": False, "order_index": 1},
        ],
    }
    q_res = client.post("/api/v1/questions", json=q_payload)
    assert q_res.status_code == 201
    question = q_res.json()
    registry["question_ids"].append(question["id"])

    exam_req = {
        "title": "Export Verification Exam",
        "exam_mode": "custom",
        "certification": unique_cert,
        "total_questions": 1,
        "passing_percentage": 70.0,
        "randomize_questions": False,
    }
    start_res = client.post("/api/v1/exams", json=exam_req)
    assert start_res.status_code == 201
    session_id = start_res.json()["id"]
    registry["session_ids"].append(session_id)

    correct_option_id = next(o["id"] for o in question["options"] if o["is_correct"])
    ans_res = client.post(
        f"/api/v1/exams/{session_id}/answer",
        json={"question_id": question["id"], "selected_option_ids": [correct_option_id]},
    )
    assert ans_res.status_code == 200

    finish_res = client.post(f"/api/v1/exams/{session_id}/finish")
    assert finish_res.status_code == 200
    completed = finish_res.json()
    assert completed["status"] == "completed"
    assert completed["score_percentage"] == 100.0

    return session_id, completed


def test_export_pdf_returns_a_valid_pdf_for_a_completed_session(cleanup_registry):
    session_id, _ = _create_completed_exam_session(cleanup_registry)

    res = client.get(f"/api/v1/export/pdf/{session_id}")

    assert res.status_code == 200
    assert res.headers["content-type"] == "application/pdf"
    assert f"exam_report_{session_id}.pdf" in res.headers["content-disposition"]
    # %PDF file signature -- proves reportlab actually produced a real PDF,
    # not just an empty or error-shaped byte string.
    assert res.content.startswith(b"%PDF")
    assert len(res.content) > 100


def test_export_excel_returns_a_workbook_with_correct_summary_and_answers(cleanup_registry):
    session_id, completed = _create_completed_exam_session(cleanup_registry)

    res = client.get(f"/api/v1/export/excel/{session_id}")

    assert res.status_code == 200
    assert res.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert f"exam_report_{session_id}.xlsx" in res.headers["content-disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    assert wb.sheetnames == ["Summary", "Answers Breakdown"]

    summary_rows = {row[0].value: row[1].value for row in wb["Summary"].iter_rows(min_row=2)}
    assert summary_rows["Title"] == "Export Verification Exam"
    assert summary_rows["Score %"] == completed["score_percentage"]
    assert summary_rows["Result"] == "passed"
    assert summary_rows["Total Questions"] == 1
    assert summary_rows["Correct Count"] == 1

    answers_sheet = wb["Answers Breakdown"]
    # Header row + exactly one answer row for this session's single question.
    assert answers_sheet.max_row == 2
    header = [c.value for c in answers_sheet[1]]
    assert header == ["Question ID", "Is Correct", "Time Spent (s)", "Confidence", "Flagged", "Bookmarked", "User Notes"]


def test_export_pdf_for_nonexistent_session_returns_404():
    res = client.get("/api/v1/export/pdf/999999")
    assert res.status_code == 404


def test_export_excel_for_nonexistent_session_returns_404():
    res = client.get("/api/v1/export/excel/999999")
    assert res.status_code == 404


def test_export_pdf_works_for_an_in_progress_session_too(cleanup_registry):
    # The export endpoints never check session.status -- an in-progress
    # session's partial report should still render, not error out.
    unique_cert = uuid.uuid4().hex
    q_payload = {
        "text": f"In-progress export test question {uuid.uuid4().hex}",
        "question_type": "single_choice",
        "certification": unique_cert,
        "options": [{"option_text": "A", "is_correct": True, "order_index": 0}],
    }
    q_res = client.post("/api/v1/questions", json=q_payload)
    cleanup_registry["question_ids"].append(q_res.json()["id"])

    start_res = client.post(
        "/api/v1/exams",
        json={"exam_mode": "custom", "certification": unique_cert, "total_questions": 1, "randomize_questions": False},
    )
    session_id = start_res.json()["id"]
    cleanup_registry["session_ids"].append(session_id)

    res = client.get(f"/api/v1/export/pdf/{session_id}")
    assert res.status_code == 200
    assert res.content.startswith(b"%PDF")
